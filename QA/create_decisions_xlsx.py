"""Generate QA decisions Excel file for hopf-layers project."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# ─── Styles ───
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
section_font = Font(name="Calibri", bold=True, size=11)
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
chosen_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
alt_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
risk_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

# ═══════════════════════════════════════════════════════════════
# SHEET 1: All QA Decisions
# ═══════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "QA Decisions"

headers = [
    "Round", "Category", "Question", "Chosen Answer", "Rationale",
    "Alternative 1", "Alt 1 Description", "Alternative 2", "Alt 2 Description",
    "Fallback Priority", "Branch"
]
col_widths = [8, 16, 45, 30, 50, 25, 40, 25, 40, 14, 12]

for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(i)].width = w

decisions = [
    # Round 1
    ["R1", "Scope", "Keep all 3 experiments or focus depth?",
     "All 3 experiments", "Broader appeal for ML venues. Variety helps workshop/preprint.",
     "Only 4.1+4.2 (physics)", "Deeper analysis, better for physics venues",
     "Only 4.2+4.3 (skip phase)", "Avoids MC bug dependency", "Low", "hl_exp1"],
    ["R1", "Infrastructure", "Fix in-house MC or use external library?",
     "Fix in-house MC code", "Full control, reproducibility. Bug likely sign convention — tractable.",
     "External lattice library", "Guaranteed correct thermalization, adds credibility",
     "Both: fix + cross-validate", "Most rigorous but most time-consuming", "Medium", "hl_mc_fix"],
    ["R1", "Strategy", "Target venue?",
     "arXiv preprint + package release", "No deadline pressure. Establish priority. Submit to venue later.",
     "Workshop paper first", "Lower bar, faster turnaround, community feedback",
     "Full conference (ICLR)", "8-10 pages, extensive ablations, 12-16 week timeline", "High", "hl_paper"],
    ["R1", "Framing", "How to frame zero-parameter layer?",
     "Feature engineering layer (like FFT)", "Non-learnable but universally used. Novelty is in WHAT features.",
     "Add learnable fiber-aware pooling", "Stronger ML story, hybrid module",
     "Preprocessing + architecture paper", "Learning happens downstream; layer provides inductive bias", "Medium", "hl_classical"],

    # Round 2
    ["R2", "Physics", "2D or 4D for topological charge?",
     "2D primary + 4D proof-of-concept", "Code works in 2D. Small 4D demo shows scaling.",
     "Stay 2D only", "Simpler, vortex winding is well-defined",
     "Extend to 4D fully", "Rigorous instanton connection, 4-6 extra weeks", "High", "hl_exp2"],
    ["R2", "Architecture", "Handle axis-angle correlation in rotation exp?",
     "Features only, no reconstruction", "Avoids correlation problem entirely.",
     "Accept as limitation", "Honest negative result if factored denoising fails",
     "Cross-attention base/fiber", "Preserves correlations, adds complexity", "Medium", "hl_exp3"],
    ["R2", "Release", "Package polish level?",
     "Full library: core + examples + docs", "Sphinx, tutorials, CI. Package IS core contribution.",
     "Minimal viable: core + tests", "Enough for pip install. Ship fast.",
     "Research code only", "Clean imports, GitHub link, no pip packaging", "High", "hl_docs"],
    ["R2", "Baselines", "S² control baseline?",
     "HopfLayer base-only ablation", "Same arch, isolates fiber contribution. No external lib.",
     "Full s2cnn spherical CNN", "Proper spherical convolutions, strongest baseline",
     "Both", "Base-only as primary + s2cnn as external reference", "Low", "hl_exp1"],

    # Round 3
    ["R3", "Numerics", "atan2 singularity handling?",
     "Clipped atan2 gradient (STE)", "Clamp to [-M, M] near singularity. Minimal code change.",
     "eps-regularized atan2", "Current approach. Simple, works for forward pass.",
     "Sigmoid-based smooth phase", "Fully smooth but forward pass approximation", "High", "hl_classical"],
    ["R3", "4D Design", "4D extension approach?",
     "Minimal: per-link HopfLayer + aggregate", "Reuses existing code. Simple.",
     "Full 4D: 4D plaquette winding", "Physically correct, substantial new code",
     "Slice-based: 2D slices of 4D", "Middle ground, some 4D structure", "Medium", "hl_exp2"],
    ["R3", "API", "Provide inverse Hopf map?",
     "Yes, exact reconstruction (S²+S¹→S³)", "Enables encode-decode. FFT analogy: forward + inverse.",
     "One-way extraction only", "Not needed for proposed experiments",
     "Provide but don't emphasize", "Utility method, not paper contribution", "Low", "hl_classical"],
    ["R3", "Parameters", "Transition temperature handling?",
     "User-tunable with principled defaults", "Constructor param. Guidance: temp ≈ a/2.",
     "Adaptive from data", "Self-tuning, no user intervention, complicates batching",
     "Make it learnable (nn.Parameter)", "ONE learnable param. Breaks pure geometric framing.", "Medium", "hl_classical"],

    # Round 4
    ["R4", "Algebra", "Octonion implementation?",
     "Cayley-Dickson from quaternions", "Reuses quaternion_multiply. Non-associativity handled by construction.",
     "Explicit 8×8 multiplication table", "Fano plane. Direct but brittle.",
     "Abstract algebra library", "Clean but heavy dependency, may not support autograd", "Low", "hl_quat"],
    ["R4", "Physics", "Phase transition model for 2D SU(2)?",
     "SU(2)+adjoint Higgs (existing model)", "2D pure SU(2) has NO phase transition. Adjoint Higgs does.",
     "3D SU(2) pure gauge", "Genuine deconfinement at β_c≈5.0. Requires 3D MC code.",
     "2D U(1) as simpler testbed", "BKT transition. But Hopf map doesn't directly apply.", "High", "hl_exp1"],
    ["R4", "Gradient", "STE smooth function choice?",
     "Clipped atan2 gradient", "Use native atan2 grad, clamp to [-M, M]. M≈100 hyperparameter.",
     "Sigmoid-based smooth phase", "φ ≈ π·sigmoid(k·a₃/a₀). Fully smooth.",
     "Detach fiber from gradient", ".detach() on fiber output. Simplest possible.", "High", "hl_classical"],
    ["R4", "Architecture", "Downstream architecture strategy?",
     "Same CNN backbone, task-specific heads", "Fair comparison. 3-layer ConvNet. Only final head varies.",
     "Task-specific architectures", "Better per-task, harder to attribute to HopfLayer",
     "Simple linear probes", "Strongest evidence for representation quality", "Medium", "hl_exp1"],

    # Round 5
    ["R5", "Scope", "Octonionic Hopf (S⁷→S¹⁵→S⁸)?",
     "Defer to v2", "Non-associativity complications. Ship 3 fibrations first.",
     "Standard Hopf invariant construction", "Per-chart. Document chart-dependence.",
     "Forward projection only, skip fiber", "Fiber is S⁷ not S¹, no phase angle.", "Low", "hl_quat"],
    ["R5", "Analysis", "Finite-size scaling depth?",
     "Validate data quality only", "Confirm β_c and ν. Main result is HopfLayer comparison.",
     "FSS as secondary contribution", "Transition signals as order parameters. Novel finding.",
     "Minimal: show phase separation exists", "Plot observable vs β. No formal FSS.", "Medium", "hl_exp1"],
    ["R5", "Compute", "GPU resources?",
     "Local RTX 4090 + cloud for 64×64", "Plenty for ≤32×32. ~60 runs total.",
     "Cloud GPU only", "A100/V100. Parallelize seeds. Costs money.",
     "CPU only", "Limits lattice sizes. MC doesn't need GPU.", "High", "hl_exp1"],

    # Round 6
    ["R6", "Narrative", "Higher Hopf fibrations in paper?",
     "Experiments on S³→S² only, Library Overview section for others",
     "Readers see framework generalizes. Experiments focus on one case.",
     "Small experiment per fibration", "Toy experiments show implementations work.",
     "Mention in conclusion only", "Cleanest narrative but under-sells library.", "Medium", "hl_paper"],
    ["R6", "Statistics", "Seeds and significance tests?",
     "5 seeds + paired t-test", "60 total runs. Standard for preprint. Mean ± std.",
     "10 seeds + Wilcoxon", "Non-parametric, 120 runs. Better for conference.",
     "3 seeds + CI only", "Minimal. 95% CI via bootstrap. Fastest.", "High", "hl_exp1"],
    ["R6", "Legal", "License?",
     "MIT", "Most permissive. Maximum adoption. Standard for ML libraries.",
     "Apache 2.0", "Patent grant. More corporate-friendly.",
     "GPL v3", "Copyleft. Limits corporate adoption.", "Low", "hl_docs"],
    ["R6", "Framing", "Include project origin story?",
     "Clean presentation, no origin story", "Purpose-built GDL tool. Save failure story for Candidate 1.",
     "Brief mention in Motivation", "One paragraph. Honest without oversharing.",
     "Full context in introduction", "Unusual but compelling research journey narrative.", "Medium", "hl_paper"],
    ["R6", "Title", "Paper title?",
     "hopf-layers: Differentiable Fiber Bundle Decompositions for GDL",
     "Broader. Matches library name. Signals multiple fibrations.",
     "Original: HopfLayer...on Gauge Fields", "Specific, emphasizes gauge theory.",
     "Fiber-Aware Neural Networks via the Hopf Fibrations", "Conceptual, plural fibrations.", "Low", "hl_paper"],
    ["R6", "Reconstruction", "Reconstruction type?",
     "Exact only, no learnable components", "Deterministic formula. FFT analogy holds.",
     "Exact + optional learnable residual", "Compensate for downstream perturbations.",
     "Learnable reconstruction network", "MLP. Loses mathematical guarantee.", "Medium", "hl_classical"],
    ["R6", "Ablation", "Ablation depth?",
     "4 configs: raw, base-only, base+fiber, full", "Clean ladder. 12 configs per experiment.",
     "2 configs: base-only vs full", "Simplest test. Fast.",
     "6+ configs: cross-combinations", "Full factorial. Most informative.", "High", "hl_exp1"],
    ["R6", "Data", "Lattice sizes?",
     "Multiple: 8, 16, 32, 64", "Gold standard for lattice studies. FSS analysis.",
     "16 to 32 only", "Standard, quick generation. 10K+ configs feasible.",
     "Up to 64", "Stronger claims, 16x slower per config.", "Medium", "hl_exp1"],
    ["R6", "Prediction", "Rotation experiment prediction target?",
     "Clean rotation matrix (9 values)", "Standard. Frobenius norm or geodesic distance metric.",
     "Euler angles (α, β, γ)", "Simpler output but gimbal lock.",
     "Axis-angle (axis + angle)", "Parallels Hopf decomposition but singularity at angle=0.", "Low", "hl_exp3"],
]

for r, row_data in enumerate(decisions, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = wrap
        cell.border = thin_border
        if c == 4:  # Chosen answer
            cell.fill = chosen_fill
        elif c in (6, 8):  # Alternatives
            cell.fill = alt_fill

ws.auto_filter.ref = f"A1:K{len(decisions)+1}"
ws.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════
# SHEET 2: Fallback Matrix
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Fallback Matrix")

fb_headers = ["Decision", "Current Choice", "Fallback Trigger", "Fallback Action", "Impact on Timeline"]
fb_widths = [35, 30, 40, 40, 20]

for i, (h, w) in enumerate(zip(fb_headers, fb_widths), 1):
    cell = ws2.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap
    cell.border = thin_border
    ws2.column_dimensions[get_column_letter(i)].width = w

fallbacks = [
    ["All 3 experiments", "All 3", "Exp 4.3 shows no improvement", "Drop 4.3, focus depth on 4.1+4.2", "Save 1-2 weeks"],
    ["Fix in-house MC", "Fix in-house", "Bug intractable after 1 week", "Switch to external library (openQCD)", "Add 1 week for integration"],
    ["arXiv + package", "arXiv preprint", "Results too strong for just preprint", "Upgrade to ICLR/NeurIPS submission", "Add 4-6 weeks for reviews"],
    ["Full library release", "Full docs+CI", "Timeline pressure", "Ship minimal viable (core+tests only)", "Save 1-2 weeks"],
    ["All 4 Hopf fibrations", "Real+Complex+Quat", "Quaternionic Hopf has numerical issues", "Ship classical only, others as 'experimental'", "Save 1 week"],
    ["Multiple lattice sizes (FSS)", "8,16,32,64", "64×64 too slow / cloud unavailable", "Drop 64, FSS with 8,16,32 only", "Save 2-3 days"],
    ["4D proof-of-concept", "Minimal 4D", "Per-link aggregation doesn't correlate", "Drop 4D, present 2D results only", "Save 1 week"],
    ["5 seeds per config", "5 seeds", "Compute budget exceeded", "3 seeds + bootstrap CI", "Save 40% compute"],
    ["Clipped atan2 STE", "STE", "Gradient instability in practice", "Fall back to eps-regularized atan2", "None"],
    ["SU(2)+adjoint Higgs", "Adjoint Higgs", "Phase transition hard to locate", "Switch to 3D SU(2) pure gauge", "Add 1-2 weeks for 3D MC code"],
    ["Same backbone across exp", "Shared CNN", "Shared backbone underperforms on all tasks", "Allow task-specific tuning", "Add 3-4 days"],
    ["Exact reconstruction only", "Exact inverse", "Downstream tasks need approximate inverse", "Add learnable residual correction", "Add 2-3 days"],
]

for r, row_data in enumerate(fallbacks, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.alignment = wrap
        cell.border = thin_border
        if c == 4:
            cell.fill = risk_fill

ws2.auto_filter.ref = f"A1:E{len(fallbacks)+1}"
ws2.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════
# SHEET 3: Project Structure
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Project Structure")

ps_headers = ["Path", "Type", "Purpose", "Status"]
ps_widths = [55, 10, 50, 12]

for i, (h, w) in enumerate(zip(ps_headers, ps_widths), 1):
    cell = ws3.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap
    cell.border = thin_border
    ws3.column_dimensions[get_column_letter(i)].width = w

structure = [
    ["hopf-layers/", "dir", "Project root", "Created"],
    ["  pyproject.toml", "file", "Build system, metadata, dependencies, pytest config", "TODO"],
    ["  README.md", "file", "Install, quick start, API overview", "TODO"],
    ["  LICENSE", "file", "MIT License", "TODO"],
    ["  .gitignore", "file", "Python, data, checkpoints exclusions", "TODO"],
    ["  src/hopf_layers/", "dir", "Main library source (src layout)", "Created"],
    ["    __init__.py", "file", "Public API: HopfLayer, QuaternionHopf, RealHopf, etc.", "TODO"],
    ["    quaternion.py", "file", "Quaternion algebra: normalize, multiply, conjugate, inverse, to/from SU(2)", "TODO"],
    ["    classical.py", "file", "Classical Hopf: S¹→S³→S² layer + HopfOutput dataclass", "TODO"],
    ["    real.py", "file", "Real Hopf: S⁰→S¹→S¹ decomposition", "TODO"],
    ["    quaternionic.py", "file", "Quaternionic Hopf: S³→S⁷→S⁴ via Cayley-Dickson", "TODO"],
    ["    transitions.py", "file", "Differentiable transition detection (tanh soft-threshold)", "TODO"],
    ["    reconstruction.py", "file", "Exact inverse: (S², S¹) → S³", "TODO"],
    ["    utils.py", "file", "Shared utilities, clipped atan2 STE", "TODO"],
    ["  tests/", "dir", "pytest test suite", "Created"],
    ["    test_quaternion.py", "file", "Quaternion algebra correctness", "TODO"],
    ["    test_classical_hopf.py", "file", "S³ constraint, S² constraint, fiber range, equivariance", "TODO"],
    ["    test_real_hopf.py", "file", "Real Hopf correctness", "TODO"],
    ["    test_quaternionic_hopf.py", "file", "S⁷/S⁴ constraints, gradient flow", "TODO"],
    ["    test_transitions.py", "file", "Winding detection: single vortex, pairs, anti-vortex", "TODO"],
    ["    test_reconstruction.py", "file", "Round-trip: decompose → reconstruct = identity", "TODO"],
    ["    test_gradient_flow.py", "file", "Backward pass through all operations, STE verification", "TODO"],
    ["  examples/", "dir", "Example scripts for users", "Created"],
    ["    01_basic_usage.py", "file", "Minimal HopfLayer usage on random quaternions", "TODO"],
    ["    02_lattice_gauge.py", "file", "Apply to SU(2) lattice gauge configuration", "TODO"],
    ["    03_rotation_features.py", "file", "Rotation denoising feature extraction", "TODO"],
    ["  docs/", "dir", "Sphinx documentation", "Created"],
    ["    conf.py", "file", "Sphinx configuration", "TODO"],
    ["    index.rst", "file", "Documentation index", "TODO"],
    ["    api/", "dir", "Auto-generated API reference", "Created"],
    ["    tutorials/", "dir", "Tutorial pages", "Created"],
    ["  experiments/", "dir", "Paper experiment code", "Created"],
    ["    exp1_phase_classification/", "dir", "Experiment 4.1: SU(2) phase classification", "Created"],
    ["      configs/", "dir", "Experiment configs (beta values, lattice sizes, seeds)", "Created"],
    ["      data/", "dir", "Generated lattice configurations (.pt)", "Created"],
    ["      models/", "dir", "Trained model checkpoints", "Created"],
    ["      results/", "dir", "Metrics, plots, statistical tests", "Created"],
    ["      run.py", "file", "Main experiment runner", "TODO"],
    ["    exp2_topological_charge/", "dir", "Experiment 4.2: topological charge", "Created"],
    ["      (same substructure as exp1)", "", "", "Created"],
    ["    exp3_rotation_denoising/", "dir", "Experiment 4.3: rotation denoising", "Created"],
    ["      (same substructure as exp1)", "", "", "Created"],
    ["    baselines/", "dir", "Shared baseline implementations", "Created"],
    ["      backbone.py", "file", "Shared 3-layer CNN backbone", "TODO"],
    ["      vanilla_cnn.py", "file", "Raw quaternion → CNN baseline", "TODO"],
    ["      quaternion_cnn.py", "file", "Quaternion-aware CNN baseline", "TODO"],
    ["    mc_generation/", "dir", "Monte Carlo configuration generation", "Created"],
    ["      su2_metropolis.py", "file", "Fixed SU(2) Metropolis MC", "TODO"],
    ["      su2_adjoint_higgs.py", "file", "SU(2)+adjoint Higgs MC", "TODO"],
    ["  paper/", "dir", "LaTeX paper source", "Created"],
    ["    main.tex", "file", "Paper manuscript", "TODO"],
    ["    figures/", "dir", "Publication-quality figures", "Created"],
    ["    references.bib", "file", "BibTeX references (10+ papers)", "TODO"],
    ["  notebooks/", "dir", "Analysis and figure generation notebooks", "Created"],
    ["    01_library_demo.ipynb", "file", "Library demonstration notebook", "TODO"],
    ["    02_experiment_analysis.ipynb", "file", "Cross-experiment analysis", "TODO"],
    ["    03_figure_generation.ipynb", "file", "Paper figure generation", "TODO"],
    ["  QA/", "dir", "Quality assurance and decision tracking", "Created"],
    ["    decisions.xlsx", "file", "This file: all QA interview decisions", "Created"],
    ["    create_decisions_xlsx.py", "file", "Script to regenerate this xlsx", "Created"],
]

for r, row_data in enumerate(structure, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.alignment = wrap
        cell.border = thin_border
        if row_data[3] == "Created":
            cell.fill = chosen_fill
        elif row_data[3] == "TODO":
            cell.fill = alt_fill

ws3.auto_filter.ref = f"A1:D{len(structure)+1}"
ws3.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════
# SHEET 4: Dependency Graph (text)
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Dependency Graph")

dg_headers = ["Branch", "Depends On", "Blocks", "Priority", "Critical Path?"]
dg_widths = [35, 40, 40, 10, 14]

for i, (h, w) in enumerate(zip(dg_headers, dg_widths), 1):
    cell = ws4.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap
    cell.border = thin_border
    ws4.column_dimensions[get_column_letter(i)].width = w

deps = [
    ["hl_pkg: Package Setup", "(none — root)", "hl_classical, hl_real, hl_quat", "HIGH", "YES"],
    ["hl_mc_fix: MC Bug Fix", "(none — root)", "hl_exp1, hl_exp2", "HIGH", "YES"],
    ["hl_classical: Classical Hopf", "hl_pkg", "hl_quat, hl_exp1, hl_exp2, hl_exp3, hl_docs", "HIGH", "YES"],
    ["hl_real: Real Hopf", "hl_pkg", "hl_docs", "MEDIUM", "No"],
    ["hl_quat: Quaternionic Hopf", "hl_pkg, hl_classical", "hl_docs", "MEDIUM", "No"],
    ["hl_exp1: Phase Classification", "hl_classical, hl_mc_fix", "hl_paper", "HIGH", "YES"],
    ["hl_exp2: Topological Charge", "hl_classical, hl_mc_fix", "hl_paper", "HIGH", "YES"],
    ["hl_exp3: Rotation Denoising", "hl_classical", "hl_paper", "MEDIUM", "No"],
    ["hl_paper: Paper Writing", "hl_exp1, hl_exp2, hl_exp3", "(none — terminal)", "HIGH", "YES"],
    ["hl_docs: Documentation", "hl_classical, hl_real, hl_quat", "(none — terminal)", "MEDIUM", "No"],
]

for r, row_data in enumerate(deps, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws4.cell(row=r, column=c, value=val)
        cell.alignment = wrap
        cell.border = thin_border
        if val == "YES":
            cell.fill = risk_fill
            cell.font = Font(bold=True)

ws4.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════
out = "C:/Users/ugure/ccode/categorical-tqft-emergence/hopf-layers/QA/decisions.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"  Sheets: {wb.sheetnames}")
print(f"  Decisions: {len(decisions)} rows")
print(f"  Fallbacks: {len(fallbacks)} rows")
print(f"  Structure: {len(structure)} rows")
print(f"  Dependencies: {len(deps)} rows")
